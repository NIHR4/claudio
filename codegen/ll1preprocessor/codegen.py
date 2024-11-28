from openpyxl import load_workbook
import re
import os

nonterminals_to_cpp_mapping = {
    "StatementList": "Symbols::STATEMENT_LIST",
    "parameter_list": "Symbols::PARAMETER_LIST",  # Assuming PARAMETER_LIST is defined
    "if_statement'": "Symbols::IF_STATEMENT_PRIME",
    "literal_expression'": "Symbols::LITERAL_EXPR",  # Assuming this corresponds
    "binary_operator": "Symbols::BINARY_OP",
    "break_statement": "Symbols::BREAK_STATEMENT",
    "increment_expression": "Symbols::INCREMENT_EXPR",
    "Statement_tail": "Symbols::STATEMENT_TAIL",
    "print_statement": "Symbols::PRINT_STATEMENT",  # Assuming PRINT_STATEMENT is defined
    "type": "Symbols::TYPE",  # Assuming TYPE is defined
    #"integer_literal": "Symbols::INTEGER_LITERAL",
    "literal_expression": "Symbols::LITERAL_EXPR",
    "term'": "Symbols::TERM_P",
    "Expression": "Symbols::EXPRESSION",
    "return_statement": "Symbols::RETURN_STATEMENT",
    "factor": "Symbols::FACTOR",
    "unary_expression": "Symbols::UNARY_EXPR",
    "while_statement": "Symbols::WHILE_STATEMENT",
    "unary_operator": "Symbols::UNARY_OP",
    "if_statement": "Symbols::IF_STATEMENT",
    "Programa": "Symbols::PROGRAM",
    "binary_expression'": "Symbols::BINARY_EXPR_P",
    "declaration_statement": "Symbols::DECLARATION_STATEMENT",  # Assuming DECLARATION_STATEMENT is defined
    "Expression_tail": "Symbols::EXPRESSION_TAIL",
    "for_statement": "Symbols::FOR_STATEMENT",
    "Statement": "Symbols::STATEMENT",
    "term": "Symbols::TERM",
    "procedure_statement": "Symbols::PROCEDURE_STATEMENT",
    "assignment_statement": "Symbols::ASSIGNMENT_STATEMENT",
    "while_statement'": "Symbols::WHILE_STATEMENT_PRIME",  # Assuming WHILE_STATEMENT_PRIME is defined
    "binary_expression": "Symbols::BINARY_EXPR",
}


terminals_to_cpp_mapping = {
    "identifier": "ID",
    "+": "PLUS",
    "-": "MINUS",
    "*": "MUL",
    "/": "DIV",
    ">": "GR_THAN",
    "<": "LS_THAN",
    "==": "EQ",
    "!=": "NEQ",
    "break": "BREAK",
    "else": "ELSE",
    "for": "FOR",
    "function": "FUNCTION",
    "if": "IF",
    "return": "RETURN",
    "while": "WHILE",
    "print": "PRINT",  # Assuming PRINT is defined; otherwise, add it to the enum
    "integer_literal": "INT_LITERAL",  # Mapping to INT_LITERAL as it corresponds to a numeric literal
    "string_literal": "STRING_LITERAL",
    "(": "LPAREN",
    ")": "RPAREN",
    "{": "LBRACE",
    "}": "RBRACE",
    ";": "SEMICOLON",
    "=": "ASSIGN",
    "++": "INCREMENT",
    "--": "DECREMENT",
    "!": "NOT",
    "int64": "INT64",
    "byte": "BYTE",
    "float": "FLOAT",
    "bool": "BOOL",
    "string": "STRING",
    "$": "DOLLAR_SIGN",  # Assuming a terminal for `$` needs to be defined
    "ε": "EPSILON",  # Assuming ε represents an empty production
}


def generate_aggregate_symbols_table(noterminals, terminals):
    m =  dict(terminals)
    m.update(noterminals)
    return sorted(m, key=lambda x: len(x[0]), reverse=True)


def remove_before_arrow(text):
    if '→' in text:
        return text.split('→', 1)[1].strip()
    return text

def convert_parse_rule(mappings, input_string):

    idx = 0
    results = []
    while idx < len(input_string):
        match_found = False
        for symbol in mappings:

            if input_string[idx] == '<':
                closingBracketIdx = input_string.find(">", idx+1)
                if closingBracketIdx != 1:
                    res = convert_parse_rule(mappings, input_string[idx+1:closingBracketIdx])
                    if len(res) > 0:
                        idx = closingBracketIdx + 1
                        match_found = True
                        results.extend(res)
                        break

            if input_string[idx:idx + len(symbol)] == symbol:
                results.append(symbol)  # Add the match to z
                idx += len(symbol)  # Skip the length of the matched substring
                match_found = True
                break  # Stop checking other symbols for this index

        if not match_found:
            idx += 1  # Move to the next character if no match is found
    return results



def generate_cpp_code(data, nonterminals, terminals):
    def to_pascal_case(text: str) -> str:
        # Split by underscores or other delimiters if necessary
        words = text.replace("'", "_Prime").split('_')
        # Capitalize each word and join them together
        pascal_case = ''.join(word.capitalize() for word in words)
        return pascal_case

    def generate_case_statements(caseElem, productions):
        cpp_cases = "\n"
        if terminals[caseElem] == 'EPSILON':
            cpp_cases += 'default:\n    return {};\n'
            return cpp_cases #early return


        cpp_cases += f"case lex::{terminals[caseElem]}:\n"
        cpp_cases += "    return {\n"  # 4 spaces for consistent C++ indentation
        for i in range(0, len(productions)):
            prod = productions[i]
            if prod in nonterminals:
                cpp_cases += f"        tag::nterm({nonterminals[prod]})"
            elif prod in terminals:
                # Edge case for empty productions
                if prod != 'ε':
                    cpp_cases += f"        tag::term(lex::{terminals[prod]})"
            if i + 1 < len(productions):
                cpp_cases += ",\n"  # Add comma only between items
            else:
                cpp_cases += "\n"
        cpp_cases += "    };\n"  # Closing bracket and semicolon
        return cpp_cases.strip()

    definitionTemplate = '''
    class {classname} : public IBaseProductionFactory {{
        virtual std::string getName() const override;
        virtual std::vector<AnySymbol> create(size_t token) override;
    }};
    '''


    identityMethodTemplate = '''
    std::string {classname}::getName() const 
    {{
        return "{classname}";
    }}
    '''

    createMethodTemplate = '''
     std::vector<AnySymbol> {classname}::create(size_t token) {{
         using lex = lexer::claudio;
         switch (token) {{
            {cases}
         }}
         throw std::runtime_error{{"no symbols matched"}};
     }}
     
     
     '''


    os.makedirs('out', exist_ok=True)
    defFiles = open("out/definitions.hpp", "w", encoding="utf-8")
    srcFiles = open("out/sources.cpp", "w", encoding="utf-8")
    for nonTerminalName, productionData in data.items():
        cname = to_pascal_case(nonTerminalName)
        print("Generating class " + cname)
        #print(definitionTemplate.format(classname=cname))
        cases = []
        defFiles.write(definitionTemplate.format(classname=cname))
        srcFiles.write(identityMethodTemplate.format(classname=cname))
        for terminalSymbol, productions in productionData.items():
            cases.append(generate_case_statements(terminalSymbol, productions))
        srcFiles.write(createMethodTemplate.format(classname=cname, cases="".join(cases)))
    defFiles.close()
    srcFiles.close()










workbook = load_workbook("C:\\Users\\pault\\Desktop\\Compilador\\test1\\ververyfinalparsingtablefrfr2.xlsx")  # Replace with your file name
sheet = workbook["Parsing Table"]  # Replace with your sheet name

# Define the range to extract from

start_cell = "B2"
end_cell = "AI30"


range_to_extract = sheet[start_cell:end_cell]

# Extract non-empty cells and flatten to a list
non_empty_cells = [cell.value for row in range_to_extract for cell in row if cell.value not in ('-', None)]

# terminals and nonterminals grouped
merged_mappings = generate_aggregate_symbols_table(nonterminals_to_cpp_mapping, terminals_to_cpp_mapping)

rules = dict()

for row_index, row in enumerate(sheet.iter_rows(min_row=2, max_row=30, min_col=1, max_col=29), start=2):
    curr_nonterminal = sheet.cell(row=row_index, column=1).value

    print('generating parsing rules for: ' + curr_nonterminal)

    for col_index, cell in enumerate(row, start=2):  # Enumerate cells in the row
        curr_terminal = sheet.cell(row=1, column=col_index)
        # Get the cell in column `X` where `X` equals the current row
        target_cell = sheet.cell(row=row_index, column=col_index)  # Column `X` corresponds to the row index
        if target_cell.value != '-' and target_cell.value != None:
            print(target_cell.value)
            if curr_nonterminal not in rules:
                rules[curr_nonterminal] = dict()
            production = convert_parse_rule(merged_mappings, remove_before_arrow(target_cell.value))
            rules[curr_nonterminal][curr_terminal.value] = production
    print("==================")

print("Codegen step...")
#print(rules)
generate_cpp_code(rules,nonterminals_to_cpp_mapping, terminals_to_cpp_mapping)